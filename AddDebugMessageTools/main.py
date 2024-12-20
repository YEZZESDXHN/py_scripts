import os
import sys
from openpyxl import load_workbook  # 导入 openpyxl
# import pandas as pd
import cantools
from PyQt5.QtCore import QCoreApplication, Qt
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QMessageBox
from addDebugMessageUI import Ui_MainWindow

def str_to_value_table(input_string):
    """将形如 "0:off 1:on 2:error" 的字符串转换为字典。"""
    value_table = {}
    try:
        pairs = input_string.split()
        for pair in pairs:
            key, value = pair.split(":")
            value_table[int(key)] = value
    except ValueError:
        # print("输入字符串格式错误，请确保格式为 'key:value key:value ...'")
        return None  # 或返回空字典 {}，取决于你的需求
    return value_table


class MainWindows(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()

        self.setupUi(self)
        self.db = None
        self.dbc_path = None
        self.new_dbc_path = None
        self.excel_df = None
        self.excel_path = None
        self.message = None
        self.load_files_state=0x00

        self.init()

    def init(self):
        self.toolButton_DB.clicked.connect(self.openDB)
        self.toolButton_excel.clicked.connect(self.openEXCEL)
        self.lineEdit_DB.textChanged.connect(self.load_dbc)
        self.lineEdit_excel.textChanged.connect(self.load_excel)
        self.pushButton_generate.clicked.connect(self.generate_new_dbc)

        self.pushButton_generate.setDisabled(1)
        self.lineEdit_DB.setReadOnly(1)
        self.lineEdit_excel.setReadOnly(1)

        self.groupBox_DB.setAcceptDrops(True)
        self.groupBox_DB.dragEnterEvent = self.dragEnterEvent
        self.groupBox_DB.dragMoveEvent = self.dragMoveEvent
        self.groupBox_DB.dropEvent = self.dropEvent_DB

        self.groupBox_excel.setAcceptDrops(True)
        self.groupBox_excel.dragEnterEvent = self.dragEnterEvent
        self.groupBox_excel.dragMoveEvent = self.dragMoveEvent
        self.groupBox_excel.dropEvent = self.dropEvent_excel

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent_DB(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            urls = event.mimeData().urls()
            file_path = urls[0].toLocalFile()
            self.lineEdit_DB.setText(file_path)
            self.dbc_path = file_path
        else:
            event.ignore()

    def dropEvent_excel(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            urls = event.mimeData().urls()
            file_path = urls[0].toLocalFile()
            self.lineEdit_excel.setText(file_path)
            self.excel_path = file_path
        else:
            event.ignore()

    def openDB(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "选择DB文件", "", "所有文件 (*)")
        if filePath:
            self.lineEdit_DB.setText(filePath)
            self.dbc_path = filePath

    def openEXCEL(self):
        filePath, _ = QFileDialog.getOpenFileName(self, "选择EXCEL文件", "", "所有文件 (*)")
        if filePath:
            self.lineEdit_excel.setText(filePath)
            self.excel_path = filePath

    def load_dbc(self,path):
        try:
            self.db = cantools.database.load_file(path)
            # print(self.db.messages[0])
            # print(self.db.messages[0].signals[0].choices)
            # signal_dict = vars(self.db.messages[0].signals[0])
            # for name, value in signal_dict.items():
            #     print(f"{name}: {value}")
            base_dir, filename = os.path.split(path)
            name, ext = os.path.splitext(filename)
            new_filename = name + "_debug" + ext
            self.new_dbc_path = os.path.join(base_dir, new_filename)

            self.load_files_state=self.load_files_state & 0x0F
            self.load_files_state=self.load_files_state | 0xF0
            if self.load_files_state==0xFF:
                self.pushButton_generate.setDisabled(0)
        except Exception as e:
            self.new_dbc_path=None
            self.pushButton_generate.setDisabled(1)
            self.load_files_state = self.load_files_state & 0x0F
            self.load_files_state = self.load_files_state | 0x00
            QMessageBox.critical(self, "Error", f"Error reading db file: {e}")
            return


    # def load_excel(self, path):
    #     try:
    #         # 使用 *args 和 **kwargs 调用 pd.read_excel
    #
    #         self.excel_df = pd.read_excel(io=path, sheet_name='Sheet1', header=2, skiprows=0)
    #         self.load_files_state = self.load_files_state & 0xF0
    #         self.load_files_state = self.load_files_state | 0x0F
    #         if self.load_files_state==0xFF:
    #             self.pushButton_generate.setDisabled(0)
    #
    #     except Exception as e:
    #         self.pushButton_generate.setDisabled(1)
    #         self.load_files_state = self.load_files_state & 0xF0
    #         self.load_files_state = self.load_files_state | 0x00
    #         QMessageBox.critical(self, "Error", f"Error reading Excel file: {e}")
    #         return
    #
    #     try:
    #         message_id = int((pd.read_excel(path, sheet_name='Sheet1', nrows=1).iloc[0, 0])[2:], 16)
    #
    #     except Exception as e:
    #         QMessageBox.critical(self, "Error", f"Error get message id: {e}")
    #         return
    #         # print(f"Error get message id: {e}")
    #     try:
    #         sender=pd.read_excel(path, sheet_name='Sheet1', header=None, nrows=1).iloc[0, 0]
    #     except Exception as e:
    #         QMessageBox.critical(self, "Error", f"Error get sender: {e}")
    #         return
    #         # print(f"Error get sender: {e}")
    #     message_name = str(sender) + "_" + str(hex(message_id))
    #
    #
    #     # valueTable =collections.OrderedDict()
    #     try:
    #         add_signals = []
    #         for index, row in self.excel_df.iterrows():
    #             signal_name = str(row['msgName'])
    #             start = int(row['起始bit'])
    #             length = int(row['长度(bit)'])
    #             if pd.isna(row['description']):
    #                 description = ""
    #             else:
    #                 description = str(row['description'])
    #             # description = str(row['description'])
    #             valueTable_str=str(row['ValueTable'])
    #             valueTable=str_to_value_table(valueTable_str)
    #             # print(valueTable,type(valueTable))
    #             new_signal = cantools.database.Signal(signal_name,
    #                                                   start=start,  # 起始位
    #                                                   length=length,  # 信号长度
    #                                                   is_signed=False,  # 是否有符号
    #                                                   comment=description,  # 注释
    #                                                   )
    #             new_signal.choices = valueTable
    #             # new_signal._dict['value_descriptions'] = valueTable
    #             add_signals.append(new_signal)
    #         # print(dir(add_signals[0]))
    #     except Exception as e:
    #         QMessageBox.critical(self, "Error", f"generate signal error: {e}")
    #         # print(f"add signal error: {e}")
    #         return
    #
    #     message = cantools.database.Message(
    #         frame_id=message_id,  # 报文ID
    #         name=message_name,  # 报文名称
    #         is_fd=True,
    #         senders=[sender],
    #         length=64,  # 报文长度 (以字节为单位)
    #         signals=add_signals,  # 空信号列表
    #         cycle_time=20
    #     )
    #     # message.replace(senders=['NewSenderNode'])
    #
    #     self.message=message

    def load_excel(self, path):
        try:
            workbook = load_workbook(filename=path, read_only=True)
            sheet = workbook.active  # or workbook["Sheet1"] if it's not the active sheet

            # 读取 message ID
            message_id_hex = sheet.cell(row=2, column=1).value
            if message_id_hex:  # Check if the cell is not empty
                try:
                    message_id = int(message_id_hex[2:], 16)
                except ValueError:
                    QMessageBox.critical(self, "Error", "Invalid message ID format in Excel.")
                    return
            else:
                QMessageBox.critical(self, "Error", "Message ID not found in Excel.")
                return

            # 读取 sender
            sender = sheet.cell(row=1, column=1).value  # Assuming sender is also in the first row and first column
            if not sender:  # Check if the cell is not empty
                QMessageBox.critical(self, "Error", "Sender not found in Excel.")
                return
            message_name = str(sender) + "_" + str(hex(message_id))

            # 读取信号数据，从第4行开始 (header=2 表示第三行是表头，跳过前两行)
            add_signals = []
            for row in sheet.iter_rows(min_row=4):
                signal_name = str(row[0].value) if row[0].value is not None else ""  # Handle potential None values
                start = int(row[1].value) if row[1].value is not None else 0
                length = int(row[2].value) if row[2].value is not None else 0
                description = str(row[4].value) if row[3].value is not None else ""
                valueTable_str = str(row[3].value) if row[4].value is not None else ""
                valueTable = str_to_value_table(valueTable_str)

                new_signal = cantools.database.Signal(signal_name, start=start, length=length, is_signed=False,
                                                      comment=description)
                new_signal.choices = valueTable
                add_signals.append(new_signal)

            message = cantools.database.Message(
                frame_id=message_id,
                name=message_name,
                is_fd=True,  # Keep as True if your messages are CAN FD
                senders=[sender],
                length=64,  # Adjust if needed
                signals=add_signals,
                cycle_time=20  # Adjust if needed
            )

            self.message = message

            self.load_files_state = self.load_files_state & 0xF0
            self.load_files_state = self.load_files_state | 0x0F
            if self.load_files_state == 0xFF:
                self.pushButton_generate.setDisabled(0)

        except Exception as e:
            self.pushButton_generate.setDisabled(1)
            self.load_files_state = self.load_files_state & 0xF0
            self.load_files_state = self.load_files_state | 0x00
            QMessageBox.critical(self, "Error", f"Error reading Excel file: {e}")
            return

    def generate_new_dbc(self):
        self.db.messages.append(self.message)
        self.db.refresh()
        # print(self.db.nodes[0])

        try:
            cantools.database.dump_file(self.db, self.new_dbc_path)  # 将修改后的数据库保存到新文件
            QMessageBox.information(self, "INFO", f"数据库已保存\n{self.new_dbc_path}")
            # print(f"数据库已保存为{self.new_dbc_path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error saving database: {e}")
            return
            # print(f"Error saving database: {e}")
        # for signal in self.message.signals:
        #     print(signal)


if __name__ == "__main__":
    QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    app = QApplication(sys.argv)
    app.setStyle("WindowsVista")
    w = MainWindows()
    w.show()
    sys.exit(app.exec_())
